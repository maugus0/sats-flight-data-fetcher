#!/usr/bin/env python3
"""
Airlines Flight Data Fetcher
Fetch flight data for any airline from AirLabs API

EXAMPLES:
  # Fetch yesterday's SQ flights
  python fetch_flights.py --airline SQ --yesterday

  # Fetch date range for Emirates
  python fetch_flights.py --airline EK --start-date 2025-01-01 --end-date 2025-01-07

  # Fetch with specific output format
  python fetch_flights.py --airline QR --yesterday --format csv

  # Interactive mode (easiest for non-technical users)
  python fetch_flights.py

  # List all available airlines
  python fetch_flights.py --list-airlines
"""

import argparse
import csv
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tqdm import tqdm

# Load environment variables
load_dotenv()

# Constants
AIRLABS_BASE_URL = "https://airlabs.co/api/v9/schedules"
CONFIG_FILE = "airlines_config.json"
OUTPUT_DIR = "outputs"
RATE_LIMIT_DELAY = 1  # seconds between API calls (between days)
MAX_RETRIES = 3
API_PAGE_LIMIT = 100  # Max results per API request (AirLabs default)
MAX_PAGINATION_PAGES = 50  # Safety limit to prevent infinite loops
# Pagination delay can be shorter than RATE_LIMIT_DELAY because:
# 1. We're within the same logical request (fetching all pages for one day)
# 2. AirLabs rate limits are typically per-minute, allowing burst requests
PAGINATION_DELAY = 0.5  # seconds between pagination requests

# CSV/Excel field names
FIELD_NAMES = [
    "flight_number",
    "departure_airport",
    "arrival_airport",
    "scheduled_departure",
    "actual_departure",
    "scheduled_arrival",
    "actual_arrival",
    "delay_minutes",
    "flight_status",
]


def load_airlines_config() -> dict:
    """Load airline configurations from JSON file."""
    config_path = Path(__file__).parent / CONFIG_FILE
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"[ERROR] Config file not found: {config_path}")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"[ERROR] Invalid JSON in config file: {config_path}")
        sys.exit(1)


def get_airline_info(code: str, config: dict) -> dict | None:
    """Get airline info by IATA or ICAO code."""
    airlines = config.get("airlines", {})

    # Direct IATA lookup
    if code.upper() in airlines:
        return airlines[code.upper()]

    # Search by ICAO code
    for iata, info in airlines.items():
        if info.get("icao", "").upper() == code.upper():
            info["iata"] = iata
            return info

    return None


def list_airlines(config: dict) -> None:
    """Print all available airlines."""
    airlines = config.get("airlines", {})

    print("\n" + "=" * 60)
    print("AVAILABLE AIRLINES")
    print("=" * 60)
    print(f"{'IATA':<6} {'ICAO':<6} {'Name':<30} {'Country'}")
    print("-" * 60)

    for iata, info in sorted(airlines.items()):
        icao = info.get("icao", "")
        name = info.get("name", "")
        country = info.get("country", "")
        print(f"{iata:<6} {icao:<6} {name:<30} {country}")

    print("-" * 60)
    print(f"Total: {len(airlines)} airlines")
    print("=" * 60 + "\n")


def validate_date(date_str: str) -> bool:
    """Validate date format YYYY-MM-DD."""
    pattern = r"^\d{4}-\d{2}-\d{2}$"
    if not re.match(pattern, date_str):
        return False
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def validate_api_key(api_key: str) -> bool:
    """Basic validation of API key format."""
    if not api_key:
        return False
    # AirLabs keys are typically 32+ characters
    return len(api_key) >= 20 and api_key.isalnum()


def get_api_key() -> str:
    """Get API key from environment or prompt."""
    api_key = os.getenv("AIRLABS_API_KEY")

    if api_key:
        print("[INFO] Using API key from .env file")
        return api_key

    print("[WARN] No API key found in .env file")
    api_key = input("Enter your AirLabs API key: ").strip()

    if not api_key:
        print("[ERROR] API key is required")
        sys.exit(1)

    return api_key


def fetch_single_page(
    api_key: str, airline_iata: str, date: str, offset: int = 0
) -> dict | None:
    """Fetch a single page of flights with retry logic."""
    params = {
        "api_key": api_key,
        "airline_iata": airline_iata,
        "dep_date": date,
        "limit": API_PAGE_LIMIT,
        "offset": offset,
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = requests.get(AIRLABS_BASE_URL, params=params, timeout=30)
            response.raise_for_status()

            data = response.json()

            if "error" in data:
                print(f"[ERROR] API error: {data['error']}")
                return None

            return data

        except requests.exceptions.Timeout:
            if attempt < MAX_RETRIES:
                wait_time = 2**attempt
                print(
                    f"[WARN] Timeout, retrying in {wait_time}s (attempt {attempt}/{MAX_RETRIES})"
                )
                time.sleep(wait_time)
            else:
                print(f"[ERROR] Request timed out after {MAX_RETRIES} attempts")
                return None

        except requests.exceptions.HTTPError as e:
            if response.status_code == 429:  # Rate limited
                wait_time = 2**attempt
                print(f"[WARN] Rate limited, waiting {wait_time}s...")
                time.sleep(wait_time)
            else:
                print(f"[ERROR] HTTP error: {e}")
                return None

        except requests.exceptions.RequestException as e:
            print(f"[ERROR] Request failed: {e}")
            return None

    return None


def fetch_flights_for_date(
    api_key: str, airline_iata: str, date: str, verbose: bool = False
) -> dict | None:
    """Fetch ALL flights for a date with pagination support."""
    all_flights = []
    offset = 0
    page_count = 0

    while page_count < MAX_PAGINATION_PAGES:  # Safety limit to prevent infinite loops
        # Fetch current page
        data = fetch_single_page(api_key, airline_iata, date, offset)

        if data is None:
            if all_flights:
                # Log warning about partial data when pagination fails mid-way
                print(
                    f"  [WARN] {date}: Pagination failed after {page_count} pages, returning partial data"
                )
                break
            return None

        flights = data.get("response", [])
        page_count += 1

        if not flights:
            break  # No more data

        all_flights.extend(flights)

        if verbose:
            print(
                f"  [DEBUG] Page {page_count}: {len(flights)} flights (offset: {offset})"
            )

        # Check if we got all flights (fewer than limit means last page)
        if len(flights) < API_PAGE_LIMIT:
            break

        # Move to next page
        offset += API_PAGE_LIMIT
        time.sleep(PAGINATION_DELAY)  # Rate limiting between pages

    # Warn if we hit the safety limit
    if page_count >= MAX_PAGINATION_PAGES:
        print(
            f"  [WARN] {date}: Hit max pagination limit ({MAX_PAGINATION_PAGES} pages). Data may be incomplete."
        )

    # Warn if exactly at limit on a single page (might be missing data)
    if len(all_flights) == API_PAGE_LIMIT and page_count == 1:
        print(
            f"  [WARN] {date}: Exactly {API_PAGE_LIMIT} flights on single page - verify pagination working"
        )

    return {"response": all_flights, "pages_fetched": page_count}


def extract_flight_data(raw_data: dict) -> list[dict]:
    """Extract and normalize flight data from API response."""
    flights = raw_data.get("response", [])
    normalized = []

    for flight in flights:
        record = {
            "flight_number": flight.get("flight_iata", ""),
            "departure_airport": flight.get("dep_iata", ""),
            "arrival_airport": flight.get("arr_iata", ""),
            "scheduled_departure": flight.get("dep_time", ""),
            "actual_departure": flight.get("dep_actual", ""),
            "scheduled_arrival": flight.get("arr_time", ""),
            "actual_arrival": flight.get("arr_actual", ""),
            "delay_minutes": flight.get("delayed", 0) or 0,
            "flight_status": flight.get("status", "unknown"),
        }
        normalized.append(record)

    return normalized


def generate_summary(flights: list[dict]) -> dict:
    """Generate summary statistics from flight data."""
    if not flights:
        return {
            "total_flights": 0,
            "average_delay_minutes": 0,
            "on_time_percentage": 0,
            "delayed_flights": 0,
            "cancelled_flights": 0,
            "flights_by_status": {},
            "top_routes": {},
        }

    total = len(flights)
    delays = [f["delay_minutes"] for f in flights if f["delay_minutes"]]
    statuses = {}
    routes = {}

    for f in flights:
        # Count by status
        status = f["flight_status"]
        statuses[status] = statuses.get(status, 0) + 1

        # Count by route
        route = f"{f['departure_airport']} → {f['arrival_airport']}"
        routes[route] = routes.get(route, 0) + 1

    avg_delay = sum(delays) / len(delays) if delays else 0
    on_time = sum(1 for f in flights if f["delay_minutes"] < 15)

    return {
        "total_flights": total,
        "average_delay_minutes": round(avg_delay, 1),
        "on_time_percentage": round(on_time / total * 100, 1) if total else 0,
        "delayed_flights": sum(1 for f in flights if f["delay_minutes"] >= 15),
        "cancelled_flights": statuses.get("cancelled", 0),
        "flights_by_status": statuses,
        "top_routes": dict(sorted(routes.items(), key=lambda x: -x[1])[:10]),
    }


def save_checkpoint(data: dict, airline: str, date: str) -> str:
    """Save raw JSON response to checkpoint file."""
    Path(OUTPUT_DIR).mkdir(exist_ok=True)
    filename = f"{OUTPUT_DIR}/checkpoint_{airline}_{date}.json"

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    return filename


def export_to_csv(flights: list[dict], filepath: str) -> None:
    """Export flight data to CSV file."""
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_NAMES)
        writer.writeheader()
        writer.writerows(flights)
    print(f"[INFO] Exported to CSV: {filepath}")


def export_to_json(flights: list[dict], summary: dict, filepath: str) -> None:
    """Export flight data and summary to JSON file."""
    output = {"summary": summary, "flights": flights}
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2)
    print(f"[INFO] Exported to JSON: {filepath}")


def export_to_excel(
    flights: list[dict], summary: dict, airline_name: str, filepath: str
) -> None:
    """Export flight data and summary to Excel file."""
    wb = Workbook()

    # Sheet 1: Flight Data
    ws_data = wb.active
    ws_data.title = "Flight Data"

    # Header styling
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    headers = [
        "Flight",
        "From",
        "To",
        "Sched. Dep",
        "Actual Dep",
        "Sched. Arr",
        "Actual Arr",
        "Delay (min)",
        "Status",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data
    for row, flight in enumerate(flights, 2):
        ws_data.cell(row=row, column=1, value=flight["flight_number"])
        ws_data.cell(row=row, column=2, value=flight["departure_airport"])
        ws_data.cell(row=row, column=3, value=flight["arrival_airport"])
        ws_data.cell(row=row, column=4, value=flight["scheduled_departure"])
        ws_data.cell(row=row, column=5, value=flight["actual_departure"])
        ws_data.cell(row=row, column=6, value=flight["scheduled_arrival"])
        ws_data.cell(row=row, column=7, value=flight["actual_arrival"])
        ws_data.cell(row=row, column=8, value=flight["delay_minutes"])
        ws_data.cell(row=row, column=9, value=flight["flight_status"])

    # Auto-adjust column widths
    for col in ws_data.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_data.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    # Sheet 2: Summary
    ws_summary = wb.create_sheet(title="Summary")

    # Title
    ws_summary.merge_cells("A1:B1")
    title_cell = ws_summary.cell(
        row=1, column=1, value=f"{airline_name} Flight Summary"
    )
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Summary metrics
    summary_data = [
        ("Total Flights", summary["total_flights"]),
        ("Average Delay (minutes)", summary["average_delay_minutes"]),
        ("On-Time Performance", f"{summary['on_time_percentage']}%"),
        ("Delayed Flights (≥15 min)", summary["delayed_flights"]),
        ("Cancelled Flights", summary["cancelled_flights"]),
    ]

    for row, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)

    # Status breakdown
    row_offset = len(summary_data) + 5
    ws_summary.cell(row=row_offset, column=1, value="Flights by Status").font = Font(
        bold=True, size=12
    )
    for i, (status, count) in enumerate(summary["flights_by_status"].items(), 1):
        ws_summary.cell(row=row_offset + i, column=1, value=status.title())
        ws_summary.cell(row=row_offset + i, column=2, value=count)

    # Top routes
    row_offset += len(summary["flights_by_status"]) + 3
    ws_summary.cell(row=row_offset, column=1, value="Top 10 Routes").font = Font(
        bold=True, size=12
    )
    for i, (route, count) in enumerate(summary["top_routes"].items(), 1):
        ws_summary.cell(row=row_offset + i, column=1, value=route)
        ws_summary.cell(row=row_offset + i, column=2, value=count)

    # Adjust column widths
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    wb.save(filepath)
    print(f"[INFO] Exported to Excel: {filepath}")


def generate_output_filename(airline: str, format: str) -> str:
    """Generate timestamped output filename."""
    Path(OUTPUT_DIR).mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    extension = "xlsx" if format == "excel" else format
    return f"{OUTPUT_DIR}/{airline}_{timestamp}.{extension}"


def interactive_mode(config: dict) -> tuple:
    """Run interactive mode for non-technical users."""
    print("\n" + "=" * 50)
    print("AIRLINES FLIGHT DATA FETCHER")
    print("Interactive Mode")
    print("=" * 50 + "\n")

    # Airline selection
    default_airline = "SQ"
    print("Enter airline IATA code (e.g., SQ, EK, BA)")
    print("Type 'list' to see all available airlines")
    airline_input = input(f"Airline [{default_airline}]: ").strip().upper()

    if airline_input == "LIST":
        list_airlines(config)
        airline_input = input(f"Airline [{default_airline}]: ").strip().upper()

    airline = airline_input if airline_input else default_airline

    # Validate airline
    airline_info = get_airline_info(airline, config)
    if not airline_info:
        print(f"[ERROR] Unknown airline: {airline}")
        print("Use --list-airlines to see available airlines")
        sys.exit(1)

    print(f"[INFO] Selected: {airline_info['name']}")

    # Date selection
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    print("\nEnter start date (YYYY-MM-DD)")
    start_date = input(f"Start date [{yesterday}]: ").strip()
    start_date = start_date if start_date else yesterday

    if not validate_date(start_date):
        print(f"[ERROR] Invalid date format: {start_date}")
        sys.exit(1)

    end_date_input = input(f"End date [{start_date}]: ").strip()
    end_date = end_date_input if end_date_input else start_date

    if not validate_date(end_date):
        print(f"[ERROR] Invalid date format: {end_date}")
        sys.exit(1)

    # Output format
    print("\nOutput format options: csv, excel, json")
    format_input = input("Format [excel]: ").strip().lower()
    output_format = (
        format_input if format_input in ["csv", "excel", "json"] else "excel"
    )

    return airline, start_date, end_date, output_format


def fetch_date_range(
    api_key: str, airline: str, start_date: str, end_date: str, verbose: bool = False
) -> list[dict]:
    """Fetch flights for a date range with pagination and progress bar."""
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")

    total_days = (end - start).days + 1
    all_flights = []
    daily_stats = []

    print(f"\n[INFO] Fetching {total_days} day(s) of data for {airline}...")
    print("[INFO] Pagination enabled - fetching ALL flights per day")

    with tqdm(total=total_days, desc="Fetching", unit="day") as pbar:
        current = start
        while current <= end:
            date_str = current.strftime("%Y-%m-%d")

            raw_data = fetch_flights_for_date(api_key, airline, date_str, verbose)

            if raw_data:
                flights = extract_flight_data(raw_data)
                day_count = len(flights)
                pages = raw_data.get("pages_fetched", 1)
                all_flights.extend(flights)
                daily_stats.append(
                    {"date": date_str, "flights": day_count, "pages": pages}
                )
                save_checkpoint(raw_data, airline, date_str)

                # Update progress bar with flight count
                pbar.set_postfix({"flights": day_count, "pages": pages})

            else:
                daily_stats.append({"date": date_str, "flights": 0, "pages": 0})

            current += timedelta(days=1)
            pbar.update(1)

            # Rate limiting between days
            if current <= end:
                time.sleep(RATE_LIMIT_DELAY)

    # Print daily statistics summary
    if daily_stats:
        flight_counts = [d["flights"] for d in daily_stats]
        avg_flights = sum(flight_counts) / len(flight_counts) if flight_counts else 0
        min_flights = min(flight_counts)
        max_flights = max(flight_counts)
        print("\n[INFO] Daily flight statistics:")
        print(f"  Min: {min_flights} | Max: {max_flights} | Avg: {avg_flights:.0f}")

    return all_flights


def main():
    """Main entry point."""
    config = load_airlines_config()

    parser = argparse.ArgumentParser(
        description="Fetch airline flight data from AirLabs API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python fetch_flights.py --airline SQ --yesterday
  python fetch_flights.py --airline EK --start-date 2025-01-01 --end-date 2025-01-07
  python fetch_flights.py --list-airlines
  python fetch_flights.py  (interactive mode)
        """,
    )

    parser.add_argument("--airline", "-a", help="Airline IATA code (e.g., SQ, EK, BA)")
    parser.add_argument("--start-date", "-s", help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", "-e", help="End date (YYYY-MM-DD)")
    parser.add_argument(
        "--yesterday", action="store_true", help="Fetch yesterday's flights"
    )
    parser.add_argument("--last-week", action="store_true", help="Fetch last 7 days")
    parser.add_argument("--last-month", action="store_true", help="Fetch last 30 days")
    parser.add_argument(
        "--format",
        "-f",
        choices=["csv", "excel", "json"],
        default="excel",
        help="Output format (default: excel)",
    )
    parser.add_argument(
        "--list-airlines", action="store_true", help="List all available airlines"
    )
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")

    args = parser.parse_args()

    # List airlines and exit
    if args.list_airlines:
        list_airlines(config)
        sys.exit(0)

    # Determine mode: interactive or CLI
    if (
        not args.airline
        and not args.yesterday
        and not args.last_week
        and not args.last_month
    ):
        airline, start_date, end_date, output_format = interactive_mode(config)
    else:
        # CLI mode
        airline = args.airline or "SQ"
        output_format = args.format

        # Validate airline
        airline_info = get_airline_info(airline, config)
        if not airline_info:
            print(f"[ERROR] Unknown airline: {airline}")
            print("Use --list-airlines to see available airlines")
            sys.exit(1)

        # Determine dates
        today = datetime.now()
        if args.yesterday:
            start_date = (today - timedelta(days=1)).strftime("%Y-%m-%d")
            end_date = start_date
        elif args.last_week:
            start_date = (today - timedelta(days=7)).strftime("%Y-%m-%d")
            end_date = (today - timedelta(days=1)).strftime("%Y-%m-%d")
        elif args.last_month:
            start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
            end_date = (today - timedelta(days=1)).strftime("%Y-%m-%d")
        else:
            start_date = args.start_date
            end_date = args.end_date or start_date

        # Validate dates
        if not start_date:
            print(
                "[ERROR] Please provide --start-date or use --yesterday/--last-week/--last-month"
            )
            sys.exit(1)

        if not validate_date(start_date):
            print(f"[ERROR] Invalid start date format: {start_date}")
            sys.exit(1)

        if not validate_date(end_date):
            print(f"[ERROR] Invalid end date format: {end_date}")
            sys.exit(1)

    # Get airline info for display
    airline_info = get_airline_info(airline, config)
    airline_name = airline_info["name"] if airline_info else airline

    # Calculate API calls needed
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    total_days = (end_dt - start_dt).days + 1
    est_time = total_days * (RATE_LIMIT_DELAY + 1)  # ~2 seconds per call

    # Confirmation
    print("\n" + "=" * 50)
    print("FETCH CONFIGURATION")
    print("=" * 50)
    print(f"Airline:      {airline_name} ({airline})")
    print(f"Date range:   {start_date} to {end_date}")
    print(f"Total days:   {total_days}")
    print(f"API calls:    {total_days}")
    print(f"Est. time:    ~{est_time} seconds")
    print(f"Output:       {output_format.upper()}")
    print("=" * 50)

    confirm = input("\nProceed? [Y/n]: ").strip().lower()
    if confirm and confirm != "y":
        print("[INFO] Cancelled by user")
        sys.exit(0)

    # Get API key
    api_key = get_api_key()

    if not validate_api_key(api_key):
        print("[WARN] API key format looks unusual, proceeding anyway...")

    # Fetch flights (with verbose flag for pagination debug info)
    verbose = args.verbose
    flights = fetch_date_range(api_key, airline, start_date, end_date, verbose)

    if not flights:
        print("\n[WARN] No flights found for the specified criteria")
        sys.exit(0)

    # Generate summary and export
    summary = generate_summary(flights)
    output_file = generate_output_filename(airline, output_format)

    if output_format == "csv":
        export_to_csv(flights, output_file)
    elif output_format == "json":
        export_to_json(flights, summary, output_file)
    else:  # excel
        export_to_excel(flights, summary, airline_name, output_file)

    # Print summary
    print("\n" + "=" * 50)
    print("SUMMARY")
    print("=" * 50)
    print(f"Total flights:    {summary['total_flights']}")
    print(f"Average delay:    {summary['average_delay_minutes']} minutes")
    print(f"On-time rate:     {summary['on_time_percentage']}%")
    print(f"Cancelled:        {summary['cancelled_flights']}")
    print("=" * 50)
    print(f"\n[SUCCESS] Output saved to: {output_file}\n")


if __name__ == "__main__":
    main()
